import os
from io import open
import openpyxl
import math
import operator

#Direcciones de los archivos
path_excel_documentos = os.path.join(os.getcwd(),"Archivos","tf","tfs.xlsx") #Para los documentos
path_excel_query = os.path.join(os.getcwd(),"Query","tf","tfs.xlsx") #para las cuerys
#contadores para rows and cols
all_rows = 0 
all_cols = 0 
all_rows_query = 0 
all_cols_query = 0
#carga de los excel
documento = openpyxl.load_workbook(path_excel_documentos)
query = openpyxl.load_workbook(path_excel_query)
#carga de las hojas de trabajo 
sheet = documento['Sheet1']
sheet_query = query['Sheet1']
#cont rows and cols
#rows for documentos
for col in sheet.iter_cols(min_col= 1, max_col=1):
    for cell in col:
        all_rows +=1
#cols for documentos
for row in sheet.iter_rows(min_row= 1, max_row=1):
    for cell in row:
        all_cols +=1

# rows for query
for col in sheet_query.iter_cols(min_col= 1, max_col=1):
    for cell in col:
        all_rows_query +=1
# cols for query
for row in sheet_query.iter_rows(min_row= 1, max_row=1):
    for cell in row:
        all_cols_query +=1

#Lista para los docuemntos y querys
lista_documentos = []
lista_query = []
#llenamos la lista con los valores
#para los documentos
for row in sheet.iter_rows(min_row= 2, max_row=all_rows,min_col=2, values_only=True):
    lista_documentos.append(row)

#para los querys
for row in sheet_query.iter_rows(min_row= 2, max_row=all_rows_query,min_col=2, values_only=True):
    lista_query.append(row)
    #print(row)

#obtenemos los valores mayores a 0 para la query
q = 0 
for i in range(all_cols_query-1):
    if lista_query[0][i] > 0:
        q += 1

#obtenemos los valores mayores a 0 y ademas sacamos las multiplicaicones
multiplicaciones = []
lista_de_mayores_de_cero = []
for i in range(len(lista_documentos)):
    count = 0
    conut2 = 0
    for j in range(all_cols-1):
        count = count + (lista_documentos[i][j] * lista_query[0][j]) #obtenemos el valor de la multiplicacion
        if lista_documentos[i][j] > 0: # para obtener los mayores a 0 de los documentos 
            conut2 += 1
    multiplicaciones.append(count)
    lista_de_mayores_de_cero.append(conut2)

#obtenemos las raices cuadradas de los elementos mayores a 0 en la query y en los documentos
raizes=[]
raiz_q = math.sqrt(q)
for i in range(len(lista_de_mayores_de_cero)):
    raizes.append(math.sqrt(lista_de_mayores_de_cero[i]))

#obtenemos la similitud
similitud = []
for i in range(len(multiplicaciones)):
    similitud.append(multiplicaciones[i]/(raizes[i]*raiz_q))

#ayuda para crear un diccioanrio para los valores y los documentos
contador_documentos = []
for i in range(len(similitud)):
    contador_documentos.append(str(i+1)+".txt")

dic = dict(zip(contador_documentos,similitud)) #convertimos las lsitas en un diccionario

lista_final = sorted(dic.items(), key=operator.itemgetter(1), reverse =True)
for i in range(20):
    print(lista_final[i])
    print("\n")
